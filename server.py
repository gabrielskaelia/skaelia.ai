# -*- coding: utf-8 -*-
"""Interface web de l'automatisation de prospection Skaelia.

Lancement :  python server.py   puis ouvrir http://localhost:5173
Accès protégé par comptes : demande d'accès -> validation par
gabriel.praud@skaelia.com -> choix du mot de passe -> connexion.
"""
import json
import os
import sys
import threading
import traceback
from datetime import datetime
from pathlib import Path

from flask import (Flask, jsonify, redirect, render_template, request,
                   send_file, session, url_for)
from werkzeug.middleware.proxy_fix import ProxyFix

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from prospection import (auth, contacts_store, enrichment, historique, mailer,
                         nicoka, pipeline)

RACINE = Path(__file__).parent
FICHIER_CONFIG = RACINE / "config.json"

app = Flask(__name__)
app.secret_key = auth.cle_secrete()
app.permanent_session_lifetime = 60 * 60 * 24 * 30  # 30 jours
# Derrière un reverse proxy HTTPS (Caddy) : respecter les en-têtes
# X-Forwarded-* pour que les liens (emails de validation) soient bien en
# https://ai.skaelia.com
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)
# Cookie de session : "Secure" seulement en ligne (HTTPS via le tunnel).
# En local (http://localhost) il doit rester non-Secure sinon la connexion casse.
_EN_LIGNE = os.environ.get("PROSPECTION_PUBLIC") == "1"
app.config.update(SESSION_COOKIE_SAMESITE="Lax", SESSION_COOKIE_SECURE=_EN_LIGNE)

# État des recherches, PAR UTILISATEUR (email -> job)
JOBS = {}
VERROU = threading.Lock()


def _job(email=None):
    """Renvoie (en le créant au besoin) l'état de recherche du compte."""
    email = email or session["email"]
    return JOBS.setdefault(email, {
        "etat": "inactif", "logs": [], "resultats": None, "erreur": "", "titre": ""})


def _config():
    try:
        return json.loads(FICHIER_CONFIG.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _ecrire_config(config):
    FICHIER_CONFIG.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")


# ---------------------------------------------------------------- accès

ROUTES_PUBLIQUES = ("/connexion", "/api/connexion", "/api/demande-acces",
                    "/valider/", "/definir-mdp/", "/api/definir-mdp", "/static/",
                    "/connexion/google", "/connexion/google/callback")


@app.before_request
def garde_acces():
    chemin = request.path
    if any(chemin == r or chemin.startswith(r) for r in ROUTES_PUBLIQUES):
        return None
    if session.get("email"):
        return None
    if chemin.startswith("/api/"):
        return jsonify({"erreur": "Non connecté"}), 401
    return redirect(url_for("page_connexion"))


@app.get("/connexion")
def page_connexion():
    if session.get("email"):
        return redirect("/")
    return render_template("connexion.html")


@app.post("/api/connexion")
def api_connexion():
    donnees = request.get_json(force=True)
    ok, erreur = auth.verifier_connexion(donnees.get("email"), donnees.get("mot_de_passe"))
    if not ok:
        return jsonify({"erreur": erreur}), 401
    session.permanent = False  # expire à la fermeture du navigateur
    session["_frais"] = True   # garde-fou : marqueur de connexion fraîche
    session["email"] = (donnees.get("email") or "").strip().lower()
    auth.enregistrer_connexion(session["email"])
    return jsonify({"ok": True})


@app.post("/api/demande-acces")
def api_demande_acces():
    donnees = request.get_json(force=True)
    email = (donnees.get("email") or "").strip().lower()
    nom = (donnees.get("nom") or "").strip()
    ok, message = auth.creer_demande(email, nom)
    if not ok:
        return jsonify({"erreur": message}), 400

    lien = request.host_url.rstrip("/") + "/valider/" + auth.jeton_validation(email)
    corps, sujet = mailer.email_validation_admin(email, nom, lien)
    envoye, erreur_envoi = mailer.envoyer(auth.ADMIN_EMAIL, sujet, corps)
    if envoye:
        return jsonify({"ok": True, "message":
                        "Skaelia a bien reçu votre demande. Vous recevrez un accès après validation."})
    # SMTP absent ou en panne : on affiche le lien de validation à transmettre
    print(f"[ACCÈS] Demande de {email} — lien de validation : {lien}")
    return jsonify({"ok": True, "sans_email": True,
                    "message": ("Skaelia a bien reçu votre demande. "
                                "Transmettez ce lien de validation à l'administrateur Skaelia :"),
                    "lien": lien})


@app.get("/valider/<jeton>")
def page_valider(jeton):
    email, erreur = auth.valider_demande(jeton)
    if not email:
        return render_template("message.html", titre="Validation impossible",
                               message=erreur, lien_mdp=""), 400
    # Comptes Google : validés directement, aucun mot de passe à définir
    _, type_auth = auth.statut_compte(email)
    if type_auth == "google":
        return render_template("message.html", titre="Compte validé ✓",
                               message=(f"Le compte {email} est validé. La personne peut désormais "
                                        "se connecter avec le bouton « Se connecter avec Google »."),
                               lien_mdp="")
    lien_mdp = request.host_url.rstrip("/") + "/definir-mdp/" + auth.jeton_mot_de_passe(email)
    corps, sujet = mailer.email_definir_mdp(lien_mdp)
    envoye, _ = mailer.envoyer(email, sujet, corps)
    if envoye:
        return render_template("message.html", titre="Compte validé ✓",
                               message=(f"Le compte {email} est validé. Un email vient de lui être "
                                        "envoyé pour choisir son mot de passe."), lien_mdp="")
    print(f"[ACCÈS] Compte {email} validé — lien mot de passe : {lien_mdp}")
    return render_template("message.html", titre="Compte validé ✓",
                           message=(f"Le compte {email} est validé. L'envoi d'email n'étant pas "
                                    "configuré, transmettez-lui ce lien pour choisir son mot de passe :"),
                           lien_mdp=lien_mdp)


@app.get("/definir-mdp/<jeton>")
def page_definir_mdp(jeton):
    return render_template("definir_mdp.html", jeton=jeton)


@app.post("/api/definir-mdp")
def api_definir_mdp():
    donnees = request.get_json(force=True)
    email, erreur = auth.definir_mot_de_passe(donnees.get("jeton"), donnees.get("mot_de_passe"))
    if not email:
        return jsonify({"erreur": erreur}), 400
    session.permanent = False  # expire à la fermeture du navigateur
    session["_frais"] = True   # garde-fou : marqueur de connexion fraîche
    session["email"] = email
    auth.enregistrer_connexion(email)
    return jsonify({"ok": True})


@app.post("/api/deconnexion")
def api_deconnexion():
    session.clear()
    return jsonify({"ok": True})


# ---------------------------------------------------------------- Google OAuth

import secrets as _secrets
from urllib.parse import urlencode as _urlencode

from curl_cffi import requests as _crequests


def _google_config():
    return _config().get("google") or {}


def _redirect_uri_google():
    return request.host_url.rstrip("/") + "/connexion/google/callback"


@app.get("/connexion/google")
def google_connexion():
    cfg = _google_config()
    if not cfg.get("client_id"):
        return render_template("message.html", titre="Google non configuré",
                               message="La connexion Google n'est pas configurée.",
                               lien_mdp=""), 400
    etat = _secrets.token_urlsafe(24)
    session["oauth_state"] = etat
    params = {
        "client_id": cfg["client_id"],
        "redirect_uri": _redirect_uri_google(),
        "response_type": "code",
        "scope": "openid email profile",
        "state": etat,
        "access_type": "online",
        "prompt": "select_account",
    }
    return redirect("https://accounts.google.com/o/oauth2/v2/auth?" + _urlencode(params))


@app.get("/connexion/google/callback")
def google_callback():
    if request.args.get("state") != session.pop("oauth_state", None):
        return render_template("message.html", titre="Connexion refusée",
                               message="Vérification de sécurité échouée, réessayez.",
                               lien_mdp=""), 400
    code = request.args.get("code")
    if not code:
        return redirect(url_for("page_connexion"))
    cfg = _google_config()
    try:
        rep = _crequests.post("https://oauth2.googleapis.com/token", data={
            "code": code,
            "client_id": cfg["client_id"],
            "client_secret": cfg["client_secret"],
            "redirect_uri": _redirect_uri_google(),
            "grant_type": "authorization_code",
        }, timeout=20)
        rep.raise_for_status()
        acces = rep.json().get("access_token")
        infos = _crequests.get("https://www.googleapis.com/oauth2/v2/userinfo",
                               headers={"Authorization": "Bearer " + acces}, timeout=20).json()
    except Exception as e:
        return render_template("message.html", titre="Connexion Google impossible",
                               message=f"Erreur lors de l'échange avec Google : {e}",
                               lien_mdp=""), 400

    email = (infos.get("email") or "").strip().lower()
    nom = infos.get("name") or ""
    if not email:
        return render_template("message.html", titre="Connexion Google impossible",
                               message="Google n'a pas renvoyé d'adresse email.", lien_mdp=""), 400

    ok, statut = auth.valider_connexion_google(email)
    if ok:
        session.permanent = False  # expire à la fermeture du navigateur
        session["_frais"] = True   # garde-fou : marqueur de connexion fraîche
        session["email"] = email
        auth.enregistrer_connexion(email)
        return redirect("/")

    if statut in ("en_attente", "valide"):
        return render_template("message.html", titre="Accès en attente",
                               message=(f"Votre compte {email} attend la validation de Skaelia. "
                                        "Vous pourrez vous connecter dès qu'il sera validé."),
                               lien_mdp="")

    # Aucun compte : on crée une demande d'accès (auth Google) et on prévient l'admin
    auth.creer_demande(email, nom, auth="google")
    lien = request.host_url.rstrip("/") + "/valider/" + auth.jeton_validation(email)
    corps, sujet = mailer.email_validation_admin(email, nom, lien)
    envoye, _ = mailer.envoyer(auth.ADMIN_EMAIL, sujet, corps)
    if not envoye:
        print(f"[ACCÈS Google] Demande de {email} — lien de validation : {lien}")
    return render_template("message.html", titre="Demande envoyée",
                           message=(f"Skaelia a bien reçu votre demande d'accès ({email}). "
                                    "Vous recevrez l'accès après validation."),
                           lien_mdp="")


@app.post("/api/changer-mdp")
def api_changer_mdp():
    donnees = request.get_json(force=True)
    ok, erreur = auth.changer_mot_de_passe(
        session["email"], donnees.get("ancien"), donnees.get("nouveau"))
    if not ok:
        return jsonify({"erreur": erreur}), 400
    return jsonify({"ok": True})


@app.get("/api/moi")
def api_moi():
    infos = auth.infos_utilisateur(session["email"]) or {"email": session["email"], "nom": ""}
    infos["admin"] = session["email"] == auth.ADMIN_EMAIL
    infos["onboarding_a_faire"] = auth.onboarding_a_faire(session["email"])
    return jsonify(infos)


@app.post("/api/onboarding-vu")
def api_onboarding_vu():
    """L'assistant de bienvenue a été complété ou passé."""
    auth.marquer_onboarding_vu(session["email"])
    return jsonify({"ok": True})


# ---------------------------------------------------------------- pages

def _version_assets():
    """Empreinte basée sur la date de modification des fichiers statiques,
    pour forcer le rafraîchissement du cache (Cloudflare/navigateur) à chaque
    modification de app.js ou style.css."""
    dossier = RACINE / "static"
    dernier = 0
    for nom in ("app.js", "style.css"):
        f = dossier / nom
        if f.exists():
            dernier = max(dernier, int(f.stat().st_mtime))
    return str(dernier)


@app.get("/")
def accueil():
    # « frais » = connexion tout juste effectuée. Le front pose alors un marqueur
    # d'onglet ; sans lui (onglet fermé puis rouvert, ou session restaurée par le
    # navigateur), le front déconnecte automatiquement. Un simple rafraîchissement
    # conserve la session (le marqueur d'onglet survit au rechargement).
    frais = session.pop("_frais", False)
    return render_template("index.html", v=_version_assets(), frais=frais)


# ------------------------------------------------- connexion Gmail (OAuth)

def _redirect_uri_gmail():
    return request.host_url.rstrip("/") + "/connexion/gmail/callback"


@app.get("/connexion/gmail")
def gmail_connexion():
    # Route sous /connexion (publique) : on exige explicitement une session.
    if not session.get("email"):
        return redirect(url_for("page_connexion"))
    cfg = _google_config()
    if not cfg.get("client_id"):
        return render_template("message.html", titre="Google non configuré",
                               message="La connexion Google n'est pas configurée sur le serveur.",
                               lien_mdp=""), 400
    etat = _secrets.token_urlsafe(24)
    session["gmail_oauth_state"] = etat
    params = {
        "client_id": cfg["client_id"],
        "redirect_uri": _redirect_uri_gmail(),
        "response_type": "code",
        "scope": "https://www.googleapis.com/auth/gmail.send openid email",
        "state": etat,
        "access_type": "offline",   # pour obtenir un refresh_token durable
        # Toujours proposer le choix du compte (permet d'en changer)
        "prompt": "select_account consent",
    }
    return redirect("https://accounts.google.com/o/oauth2/v2/auth?" + _urlencode(params))


@app.get("/connexion/gmail/callback")
def gmail_callback():
    if not session.get("email"):
        return redirect(url_for("page_connexion"))
    if request.args.get("state") != session.pop("gmail_oauth_state", None):
        return render_template("message.html", titre="Connexion refusée",
                               message="Vérification de sécurité échouée, réessayez.",
                               lien_mdp=""), 400
    code = request.args.get("code")
    if not code:
        return redirect("/")
    cfg = _google_config()
    try:
        rep = _crequests.post("https://oauth2.googleapis.com/token", data={
            "code": code,
            "client_id": cfg["client_id"],
            "client_secret": cfg["client_secret"],
            "redirect_uri": _redirect_uri_gmail(),
            "grant_type": "authorization_code",
        }, timeout=20)
        rep.raise_for_status()
        jetons = rep.json()
        infos = _crequests.get("https://www.googleapis.com/oauth2/v2/userinfo",
                               headers={"Authorization": "Bearer " + jetons["access_token"]},
                               timeout=20).json()
    except Exception as e:
        return render_template("message.html", titre="Connexion Gmail impossible",
                               message=f"Erreur lors de l'échange avec Google : {e}",
                               lien_mdp=""), 400
    if not jetons.get("refresh_token"):
        return render_template("message.html", titre="Connexion Gmail incomplète",
                               message="Google n'a pas fourni d'autorisation durable — réessayez.",
                               lien_mdp=""), 400
    auth.ecrire_gmail_oauth(session["email"], {
        "adresse": infos.get("email", ""),
        "refresh_token": jetons["refresh_token"],
    })
    return redirect("/?gmail=ok")


# ---------------------------------------------------------------- administration

def _refus_si_pas_admin():
    if session.get("email") != auth.ADMIN_EMAIL:
        return jsonify({"erreur": "Réservé à l'administrateur."}), 403
    return None


@app.get("/api/admin/utilisateurs")
def api_admin_utilisateurs():
    refus = _refus_si_pas_admin()
    if refus:
        return refus
    return jsonify(auth.lister_utilisateurs())


@app.post("/api/admin/mdp")
def api_admin_mdp():
    refus = _refus_si_pas_admin()
    if refus:
        return refus
    donnees = request.get_json(force=True)
    ok, erreur = auth.admin_changer_mdp(donnees.get("email"), donnees.get("mot_de_passe"))
    if not ok:
        return jsonify({"erreur": erreur}), 400
    return jsonify({"ok": True})


@app.post("/api/admin/supprimer")
def api_admin_supprimer():
    refus = _refus_si_pas_admin()
    if refus:
        return refus
    donnees = request.get_json(force=True)
    email = (donnees.get("email") or "").strip().lower()
    ok, erreur = auth.supprimer_utilisateur(email)
    if not ok:
        return jsonify({"erreur": erreur}), 400
    # Nettoyage des données du compte : carnet de contacts + fichiers Excel
    contacts_store.supprimer_utilisateur(email)
    import shutil
    shutil.rmtree(historique.DOSSIER_RESULTATS / historique.slug_user(email),
                  ignore_errors=True)
    return jsonify({"ok": True})


# ---------------------------------------------------------------- réglages

@app.get("/api/reglages")
def lire_reglages():
    config = _config()
    cle = config.get("usebouncer_api_key", "")
    smtp_perso = auth.lire_smtp_perso(session["email"])
    gmail = auth.lire_gmail_oauth(session["email"])
    smtp_ok = bool(smtp_perso.get("utilisateur") and smtp_perso.get("mot_de_passe"))
    return jsonify({
        "usebouncer_configuree": bool(cle),
        "fullenrich_configuree": bool(config.get("fullenrich_api_key")),
        "exclusions_cabinets": config.get("exclusions_cabinets", []),
        "smtp_configure": mailer.smtp_configure(),
        # Connexion d'envoi propre au compte : Gmail OAuth d'abord, sinon SMTP
        "email_perso_configure": bool(gmail.get("refresh_token")) or smtp_ok,
        "email_perso_adresse": gmail.get("adresse") or smtp_perso.get("utilisateur", ""),
        "gmail_oauth": bool(gmail.get("refresh_token")),
        "email_perso_hote": smtp_perso.get("hote", ""),
        "email_perso_port": smtp_perso.get("port", 587),
    })


@app.post("/api/reglages")
def ecrire_reglages():
    donnees = request.get_json(force=True)
    config = _config()
    if donnees.get("usebouncer_api_key"):
        config["usebouncer_api_key"] = donnees["usebouncer_api_key"].strip()
    if donnees.get("fullenrich_api_key"):
        config["fullenrich_api_key"] = donnees["fullenrich_api_key"].strip()
    if "exclusions_cabinets" in donnees:
        config["exclusions_cabinets"] = [
            e.strip() for e in donnees["exclusions_cabinets"] if e.strip()]
    if "smtp" in donnees:
        smtp = donnees["smtp"] or {}
        existant = config.get("smtp") or {}
        config["smtp"] = {
            "hote": (smtp.get("hote") or "").strip(),
            "port": int(smtp.get("port") or 587),
            "utilisateur": (smtp.get("utilisateur") or "").strip(),
            "mot_de_passe": (smtp.get("mot_de_passe") or "").strip()
                            or existant.get("mot_de_passe", ""),
            "expediteur": (smtp.get("utilisateur") or "").strip(),
        }
    _ecrire_config(config)
    # Connexion Gmail/SMTP personnelle (stockée sur le compte, pas en global)
    if "smtp_perso" in donnees:
        auth.ecrire_smtp_perso(session["email"], donnees["smtp_perso"] or {})
    if "gmail_oauth" in donnees and not donnees["gmail_oauth"]:
        auth.ecrire_gmail_oauth(session["email"], {})   # déconnexion Gmail
    return jsonify({"ok": True})


@app.post("/api/envoyer-email")
def api_envoyer_email():
    """Envoie un email de prospection via la connexion Gmail du compte."""
    donnees = request.get_json(force=True)
    destinataire = (donnees.get("a") or "").strip()
    sujet = (donnees.get("sujet") or "").strip() or "Prise de contact — Skaelia"
    corps = donnees.get("corps") or ""
    if "@" not in destinataire:
        return jsonify({"erreur": "Destinataire invalide."}), 400
    if not corps.strip():
        return jsonify({"erreur": "Le message est vide."}), 400
    gmail = auth.lire_gmail_oauth(session["email"])
    if gmail.get("refresh_token"):
        ok, erreur = mailer.envoyer_gmail(gmail, destinataire, sujet, corps)
    else:
        conf = auth.lire_smtp_perso(session["email"])
        if not (conf.get("utilisateur") and conf.get("mot_de_passe")):
            return jsonify({"erreur": "Connexion Gmail non configurée (Réglages)."}), 400
        ok, erreur = mailer.envoyer_pour(conf, destinataire, sujet, corps)
    if not ok:
        return jsonify({"erreur": erreur}), 502
    return jsonify({"ok": True})


# ---------------------------------------------------------------- exécution

def _executer_en_fond(email, params):
    job = _job(email)

    def log(message):
        job["logs"].append({"heure": datetime.now().strftime("%H:%M:%S"),
                            "texte": str(message)})
    try:
        job["resultats"] = pipeline.executer(params, log=log)
        job["etat"] = "termine"
    except Exception as e:
        job["erreur"] = str(e)
        job["etat"] = "erreur"
        log(f"Erreur : {e}")
        traceback.print_exc()


@app.post("/api/lancer")
def lancer():
    job = _job()
    with VERROU:
        if job["etat"] == "en_cours":
            return jsonify({"erreur": "Une recherche est déjà en cours"}), 409
        donnees = request.get_json(force=True)
        # Poste et secteur sont tous deux facultatifs, mais il en faut au moins un
        # (le secteur sert de mot-clé de recherche à défaut de poste précis).
        poste = (donnees.get("poste") or "").strip() or (donnees.get("secteur") or "").strip()
        if not poste:
            return jsonify({"erreur": "Indiquez un poste ou un secteur"}), 400

        config = _config()
        lieu = (donnees.get("lieu") or "").strip() or (donnees.get("region") or "").strip()
        # Exclure du sourcing : cabinets concurrents / agences d'intérim.
        # Les entreprises clientes (Nicoka) ne sont PLUS exclues : elles sont
        # classées « client » (les autres « prospect »).
        params = {
            "poste": poste,
            "lieu": lieu,
            "exclusions": list(config.get("exclusions_cabinets", [])),
            "clients": nicoka.liste_clients(),
            "verifier_emails": False,
            "usebouncer_api_key": "",
            # Résultats et « nouveautés » isolés par compte
            "dossier_resultats": str(historique.dossier_user(session["email"])),
            "cle_user": historique.slug_user(session["email"]),
        }
        types = donnees.get("types_entreprise")
        if isinstance(types, list) and types:
            params["types_entreprise"] = [t for t in types if t in ("prospect", "client")]
        if isinstance(donnees.get("contrats"), list):
            params["contrats"] = donnees["contrats"]
        if isinstance(donnees.get("sources"), list) and donnees["sources"]:
            params["sources"] = donnees["sources"]
        for cle in ("anciennete_jours", "pages", "max_entreprises", "nb_contacts_cible"):
            if donnees.get(cle) not in (None, ""):
                params[cle] = int(donnees[cle])
        if donnees.get("teletravail_uniquement"):
            params["teletravail_uniquement"] = True
        if donnees.get("recherche_amelioree"):
            params["recherche_amelioree"] = True
        if donnees.get("inclure_cabinets"):
            params["inclure_cabinets"] = True
        job.update({"etat": "en_cours", "logs": [], "resultats": None,
                    "erreur": "", "titre": poste + (f" — {params['lieu']}" if params["lieu"] else "")})
        threading.Thread(target=_executer_en_fond,
                         args=(session["email"], params), daemon=True).start()
    return jsonify({"ok": True})


@app.get("/api/statut")
def statut():
    job = _job()
    reponse = {"etat": job["etat"], "titre": job["titre"],
               "logs": job["logs"], "erreur": job["erreur"]}
    if job["etat"] == "termine" and job["resultats"]:
        r = job["resultats"]
        reponse["synthese"] = {
            "nb_offres": len(r["offres"]),
            "nb_entreprises": len(r["entreprises"]),
            "nb_contacts": len(r["contacts"]),
            "nb_nouvelles": r["nb_nouvelles"],
        }
    return jsonify(reponse)


@app.get("/api/resultats")
def resultats():
    job = _job()
    if job["etat"] != "termine" or not job["resultats"]:
        return jsonify({"erreur": "Aucun résultat disponible"}), 404
    r = job["resultats"]
    return jsonify({
        "offres": r["offres"],
        "entreprises": r["entreprises"],
        "contacts": r["contacts"],
        "nb_nouvelles": r["nb_nouvelles"],
        "fichier": Path(r["fichier"]).name if r["fichier"] else "",
    })


@app.get("/api/telecharger")
def telecharger():
    job = _job()
    if not (job["resultats"] and job["resultats"].get("fichier")):
        return jsonify({"erreur": "Aucun fichier disponible"}), 404
    chemin = Path(job["resultats"]["fichier"])
    if not chemin.exists():
        return jsonify({"erreur": "Fichier introuvable"}), 404
    return send_file(chemin, as_attachment=True, download_name=chemin.name)


# ---------------------------------------------------------------- historique (par compte)

@app.get("/api/historique")
def api_historique():
    return jsonify(historique.lister(session["email"]))


@app.delete("/api/historique")
def api_historique_vider():
    n = historique.vider(session["email"])
    return jsonify({"ok": True, "supprimes": n})


@app.delete("/api/historique/<nom_fichier>")
def api_historique_supprimer(nom_fichier):
    ok = historique.supprimer(session["email"], nom_fichier)
    return jsonify({"ok": ok})


@app.get("/api/historique/<nom_fichier>")
def api_historique_charger(nom_fichier):
    donnees = historique.charger(session["email"], nom_fichier)
    if not donnees:
        return jsonify({"erreur": "Recherche introuvable"}), 404
    # On la place comme résultat courant DU COMPTE (pour le téléchargement Excel)
    job = _job()
    job["resultats"] = donnees
    job["etat"] = "termine"
    return jsonify({
        "offres": donnees["offres"],
        "entreprises": donnees["entreprises"],
        "contacts": donnees["contacts"],
        "nb_nouvelles": donnees["nb_nouvelles"],
        "fichier": Path(donnees["fichier"]).name,
    })


# ---------------------------------------------------------------- carnet de contacts

@app.get("/api/mes-contacts")
def api_mes_contacts():
    return jsonify(contacts_store.lister(session["email"]))


@app.post("/api/mes-contacts")
def api_ajouter_contacts():
    donnees = request.get_json(force=True)
    contacts = donnees.get("contacts") or []
    if not isinstance(contacts, list):
        contacts = [contacts]
    # Vérification Nicoka faite à l'ajout : on attache le statut à chaque contact
    # (persisté) et on signale ceux déjà contactés sous N jours.
    recents = []
    cache = nicoka.charger_cache()
    for c in contacts:
        if cache:
            st = nicoka.statut_pour(c.get("email", ""), c.get("nom", ""), cache)
            c["nicoka"] = {
                "verifie": True,
                "en_base": st.get("en_base", False),
                "recent": st.get("recent", False),
                "jours": st.get("jours_depuis_action"),
            }
            if st.get("recent"):
                recents.append({"nom": c.get("nom", ""),
                                "jours": st.get("jours_depuis_action")})
        else:
            c["nicoka"] = {"verifie": False}
        # PAS d'enrichissement automatique ici : les crédits FullEnrich ne sont
        # consommés qu'à la demande, au clic dans « Prendre contact manuellement »
        # (voir /api/mes-contacts/enrichir).
    ajoutes, liste = contacts_store.ajouter(session["email"], contacts)

    return jsonify({"ok": True, "ajoutes": ajoutes, "contacts": liste,
                    "recents_nicoka": recents, "jours": nicoka.jours_prospection(),
                    "nicoka_ok": bool(cache)})


@app.post("/api/mes-contacts/enrichir")
def api_enrichir_contact():
    """Enrichissement FullEnrich À LA DEMANDE (consomme des crédits), UN champ à
    la fois : `champ` = "email" (~1 crédit) ou "telephone" (~10 crédits).
    Déclenché au clic dans « Prendre contact manuellement »."""
    if not enrichment.est_configure():
        return jsonify({"erreur": "Enrichissement indisponible : ajoute ta clé "
                                  "FullEnrich dans les Réglages."}), 400
    donnees = request.get_json(force=True) or {}
    cle = donnees.get("cle", "")
    champ = donnees.get("champ", "")
    if champ not in ("email", "telephone"):
        return jsonify({"erreur": "Champ à enrichir invalide."}), 400
    contact = next((c for c in contacts_store.lister(session["email"])
                    if contacts_store.cle_contact(c) == cle), None)
    if not contact:
        return jsonify({"erreur": "Contact introuvable."}), 404

    resultats = enrichment.enrichir_lot([contact], champs=(champ,))
    r = (resultats[0] if resultats else {}) or {}
    if champ == "email":
        maj = {"email_recherche": "fait"}
        if r.get("email"):
            maj["email"] = r["email"]
            maj["statut_email"] = r.get("statut_email", "")
        trouve = bool(r.get("email"))
    else:
        maj = {"tel_recherche": "fait"}
        if r.get("telephone"):
            maj["telephone"] = r["telephone"]
        trouve = bool(r.get("telephone"))
    contact_maj = contacts_store.mettre_a_jour(session["email"], cle, maj)
    return jsonify({"ok": True, "contact": contact_maj, "trouve": trouve})


@app.post("/api/mes-contacts/verifier-nicoka")
def api_verifier_nicoka_tous():
    cache = nicoka.charger_cache()
    if not cache:
        return jsonify({"erreur": "Base Nicoka non synchronisée (voir Réglages)."}), 400

    def calcul(c):
        st = nicoka.statut_pour(c.get("email", ""), c.get("nom", ""), cache)
        return {"verifie": True, "en_base": st.get("en_base", False),
                "recent": st.get("recent", False), "jours": st.get("jours_depuis_action")}

    liste = contacts_store.maj_nicoka_tous(session["email"], calcul)
    deja = sum(1 for c in liste if (c.get("nicoka") or {}).get("recent"))
    return jsonify({"ok": True, "contacts": liste,
                    "total": len(liste), "deja_contactes": deja,
                    "non_prospectes": len(liste) - deja,
                    "jours": nicoka.jours_prospection()})


@app.post("/api/verifier-email")
def api_verifier_email():
    donnees = request.get_json(force=True)
    email = (donnees.get("email") or "").strip()
    if not email:
        return jsonify({"erreur": "Aucune adresse à vérifier"}), 400
    cle_api = _config().get("usebouncer_api_key", "")
    if not cle_api:
        return jsonify({"erreur": "Clé UseBouncer non configurée"}), 400
    from prospection import emails as _emails
    statut, raison = _emails.verifier_email(email, cle_api)
    # Mémoriser le statut sur le contact du carnet si une clé est fournie
    if donnees.get("cle"):
        contacts_store.mettre_a_jour(session["email"], donnees["cle"],
                                     {"statut_email": statut})
    return jsonify({"ok": True, "email": email, "statut": statut, "raison": raison})


@app.post("/api/mes-contacts/supprimer")
def api_supprimer_contacts():
    donnees = request.get_json(force=True)
    cles = donnees.get("cles") or ([donnees["cle"]] if donnees.get("cle") else [])
    liste = contacts_store.supprimer(session["email"], cles)
    return jsonify({"ok": True, "contacts": liste})


@app.post("/api/mes-contacts/maj")
def api_maj_contact():
    donnees = request.get_json(force=True)
    contact = contacts_store.mettre_a_jour(session["email"], donnees.get("cle"), donnees)
    if not contact:
        return jsonify({"erreur": "Contact introuvable"}), 404
    return jsonify({"ok": True, "contact": contact})


@app.get("/api/nicoka/etat")
def api_nicoka_etat():
    cache = nicoka.charger_cache()
    return jsonify({
        "configure": nicoka.est_configure(),
        "sync_en_cours": nicoka.SYNC["en_cours"],
        "synced_at": cache.get("synced_at") if cache else None,
        "total": cache.get("total") if cache else 0,
        "jours": nicoka.jours_prospection(),
    })


@app.post("/api/nicoka/sync")
def api_nicoka_sync():
    if not nicoka.est_configure():
        return jsonify({"erreur": "Nicoka n'est pas configuré"}), 400
    if nicoka.SYNC["en_cours"]:
        return jsonify({"ok": True, "deja": True})

    def _sync():
        nicoka.synchroniser(log=lambda m: None)
        try:
            nicoka.synchroniser_references()  # références clients pour les messages
        except Exception as e:
            print("Sync références Nicoka:", e)

    threading.Thread(target=_sync, daemon=True).start()
    return jsonify({"ok": True})


def _auto_sync_nicoka():
    """Synchronise Nicoka en tâche de fond si le cache est absent ou date de
    plus de 24 h. Appelé au démarrage : plus besoin de bouton dans l'interface."""
    if not nicoka.est_configure() or nicoka.SYNC["en_cours"]:
        return
    from datetime import timedelta
    cache = nicoka.charger_cache()
    frais = False
    if cache and cache.get("synced_at"):
        try:
            vu = datetime.fromisoformat(cache["synced_at"])
            frais = (datetime.now() - vu) < timedelta(hours=24)
        except ValueError:
            frais = False
    if frais:
        return

    def _sync():
        try:
            nicoka.synchroniser(log=lambda m: None)
            nicoka.synchroniser_references()
        except Exception as e:
            print("Auto-sync Nicoka:", e)

    threading.Thread(target=_sync, daemon=True).start()


@app.get("/api/references")
def api_references():
    return jsonify(nicoka.references_pour(request.args.get("poste", ""), n=2))


@app.get("/api/nicoka/sync-statut")
def api_nicoka_sync_statut():
    return jsonify({
        "en_cours": nicoka.SYNC["en_cours"],
        "message": nicoka.SYNC["message"],
        "total": nicoka.SYNC["total"],
        "recus": nicoka.SYNC["recus"],
        "erreur": nicoka.SYNC["erreur"],
    })


@app.get("/api/nicoka/contacts")
def api_nicoka_contacts():
    return jsonify(nicoka.lister(
        recherche=request.args.get("recherche", ""),
        possible_prospection=request.args.get("prospection") == "1",
        page=int(request.args.get("page", 1)),
    ))


@app.get("/api/mes-contacts/export")
def api_export_contacts():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    contacts = contacts_store.lister(session["email"])
    wb = Workbook()
    ws = wb.active
    ws.title = "Mes contacts"
    entetes = ["Nom", "Poste", "Entreprise", "Email", "Statut email",
               "Téléphone", "Profil LinkedIn", "Ajouté le"]
    for i, titre in enumerate(entetes, start=1):
        c = ws.cell(row=1, column=i, value=titre)
        c.fill = PatternFill("solid", fgColor="3D4451")
        c.font = Font(color="FFFFFF", bold=True)
    for r, ct in enumerate(contacts, start=2):
        for i, cle in enumerate(["nom", "poste", "entreprise", "email", "statut_email",
                                 "telephone", "url_linkedin", "ajoute_le"], start=1):
            ws.cell(row=r, column=i, value=ct.get(cle, ""))
    for col, largeur in zip("ABCDEFGH", [26, 34, 26, 30, 16, 18, 50, 20]):
        ws.column_dimensions[col].width = largeur
    ws.freeze_panes = "A2"
    flux = io.BytesIO()
    wb.save(flux)
    flux.seek(0)
    return send_file(flux, as_attachment=True, download_name="mes_contacts.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


try:
    _auto_sync_nicoka()
except Exception as _e:
    print("Auto-sync Nicoka au démarrage:", _e)


if __name__ == "__main__":
    print("Interface disponible sur http://localhost:5173")
    app.run(host="127.0.0.1", port=5173, debug=False)
