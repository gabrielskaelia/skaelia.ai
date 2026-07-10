# -*- coding: utf-8 -*-
"""Historique des recherches : relit les fichiers Excel du dossier resultats/
pour ré-afficher les recherches passées dans l'interface, sans tout relancer.

Le nom des fichiers suit le format :
    prospection_<slug>_<AAAA-MM-JJ>_<HHhMM>.xlsx
"""
import hashlib
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

RACINE = Path(__file__).parent.parent
DOSSIER_RESULTATS = RACINE / "resultats"


def slug_user(email):
    """Identifiant de dossier stable et sûr, propre à chaque compte."""
    email = (email or "anonyme").strip().lower()
    base = unicodedata.normalize("NFKD", email)
    base = "".join(c for c in base if not unicodedata.combining(c))
    base = re.sub(r"[^a-z0-9]+", "-", base).strip("-")[:40] or "user"
    # Suffixe de hash pour éviter toute collision
    h = hashlib.sha1(email.encode("utf-8")).hexdigest()[:8]
    return f"{base}-{h}"


def dossier_user(email):
    """Dossier de résultats propre au compte (créé si besoin)."""
    d = DOSSIER_RESULTATS / slug_user(email)
    d.mkdir(parents=True, exist_ok=True)
    return d

# Correspondance en-tête Excel -> clé interne, par onglet
_COLONNES = {
    "Offres": {
        "Nouveau": "nouveau", "Titre du poste": "titre", "Entreprise": "entreprise",
        "Lieu": "lieu", "Contrat": "contrat", "Salaire": "salaire",
        "Publiée": "date", "Source": "source", "Lien": "url",
    },
    "Entreprises": {
        "Entreprise": "entreprise", "Nb d'offres": "nb_offres",
        "Postes recherchés": "postes", "Lieux": "lieux",
        "Sources": "sources", "Domaine email": "domaine",
    },
    "Contacts": {
        "Entreprise": "entreprise", "Nom": "nom", "Poste": "poste",
        "Profil LinkedIn": "url_linkedin", "Email": "email",
        "Statut email": "statut_email", "Extrait profil": "extrait",
    },
}


def _lire_feuille(ws, correspondance):
    lignes = list(ws.iter_rows(values_only=True))
    if not lignes:
        return []
    entetes = [str(c) if c is not None else "" for c in lignes[0]]
    resultat = []
    for ligne in lignes[1:]:
        if all(v is None or v == "" for v in ligne):
            continue
        objet = {}
        for i, entete in enumerate(entetes):
            cle = correspondance.get(entete)
            if cle:
                valeur = ligne[i] if i < len(ligne) else ""
                objet[cle] = "" if valeur is None else valeur
        # Champs listes (affichés séparés par des barres dans l'Excel)
        for champ in ("postes", "lieux"):
            if isinstance(objet.get(champ), str) and " | " in objet[champ]:
                objet[champ] = [x.strip() for x in objet[champ].split("|")]
        resultat.append(objet)
    return resultat


def _titre_depuis_nom(nom_fichier):
    """prospection_assistant-commercial-angers_2026-07-08_16h41 -> "assistant commercial angers"."""
    base = re.sub(r"^prospection_", "", nom_fichier)
    base = re.sub(r"_\d{4}-\d{2}-\d{2}_\d{2}h\d{2}$", "", base)
    return base.replace("-", " ").strip()


def _date_depuis_nom(nom_fichier):
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})_(\d{2})h(\d{2})", nom_fichier)
    if not m:
        return ""
    a, mo, j, h, mi = m.groups()
    return f"{j}/{mo}/{a} à {h}h{mi}"


def lister(email):
    """Liste des recherches passées DE CE COMPTE, la plus récente d'abord."""
    dossier = DOSSIER_RESULTATS / slug_user(email)
    if not dossier.exists():
        return []
    fichiers = sorted(
        dossier.glob("prospection_*.xlsx"),
        key=lambda p: p.stat().st_mtime, reverse=True,
    )
    entrees = []
    for f in fichiers:
        try:
            wb = load_workbook(f, read_only=True)
            nb = {}
            for onglet in ("Offres", "Entreprises", "Contacts"):
                nb[onglet] = (wb[onglet].max_row - 1) if onglet in wb.sheetnames else 0
            wb.close()
        except Exception:
            continue
        entrees.append({
            "fichier": f.name,
            "titre": _titre_depuis_nom(f.stem),
            "date": _date_depuis_nom(f.stem),
            "nb_offres": max(nb["Offres"], 0),
            "nb_entreprises": max(nb["Entreprises"], 0),
            "nb_contacts": max(nb["Contacts"], 0),
        })
    return entrees


def _chemin_sur(email, nom_fichier):
    """Chemin sécurisé d'un fichier de résultats DU COMPTE (anti-traversée)."""
    if "/" in nom_fichier or "\\" in nom_fichier or ".." in nom_fichier:
        return None
    chemin = DOSSIER_RESULTATS / slug_user(email) / nom_fichier
    if chemin.suffix.lower() != ".xlsx" or not chemin.exists():
        return None
    return chemin


def supprimer(email, nom_fichier):
    """Supprime un fichier de résultats du compte. Retourne True si supprimé."""
    chemin = _chemin_sur(email, nom_fichier)
    if not chemin:
        return False
    chemin.unlink()
    return True


def vider(email):
    """Supprime toutes les recherches du compte. Retourne le nombre supprimé."""
    dossier = DOSSIER_RESULTATS / slug_user(email)
    if not dossier.exists():
        return 0
    n = 0
    for f in dossier.glob("prospection_*.xlsx"):
        try:
            f.unlink()
            n += 1
        except OSError:
            pass
    return n


def charger(email, nom_fichier):
    """Relit un fichier Excel de résultats. Retourne le dict de résultats
    (offres, entreprises, contacts, fichier) ou None si introuvable/invalide."""
    # Sécurité : n'accepter qu'un fichier du dossier du compte
    if "/" in nom_fichier or "\\" in nom_fichier or ".." in nom_fichier:
        return None
    chemin = DOSSIER_RESULTATS / slug_user(email) / nom_fichier
    if not chemin.exists() or chemin.suffix.lower() != ".xlsx":
        return None
    try:
        wb = load_workbook(chemin, read_only=True)
    except Exception:
        return None
    donnees = {}
    for onglet, cle in (("Offres", "offres"), ("Entreprises", "entreprises"),
                        ("Contacts", "contacts")):
        donnees[cle] = _lire_feuille(wb[onglet], _COLONNES[onglet]) if onglet in wb.sheetnames else []
    wb.close()
    donnees["nb_nouvelles"] = sum(1 for o in donnees["offres"] if o.get("nouveau"))
    donnees["fichier"] = str(chemin)
    return donnees
