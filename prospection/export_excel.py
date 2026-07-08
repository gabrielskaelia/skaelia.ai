# -*- coding: utf-8 -*-
"""Export des résultats dans un classeur Excel à trois onglets :
Offres, Entreprises, Contacts."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ENTETE_FILL = PatternFill("solid", fgColor="1F4E79")
ENTETE_FONT = Font(color="FFFFFF", bold=True)


def _ecrire_feuille(ws, colonnes, lignes, largeurs):
    for i, (titre, _) in enumerate(colonnes, start=1):
        cellule = ws.cell(row=1, column=i, value=titre)
        cellule.fill = ENTETE_FILL
        cellule.font = ENTETE_FONT
        cellule.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = largeurs[i - 1]
    for r, ligne in enumerate(lignes, start=2):
        for c, (_, cle) in enumerate(colonnes, start=1):
            valeur = ligne.get(cle, "")
            if isinstance(valeur, list):
                valeur = " | ".join(valeur)
            cellule = ws.cell(row=r, column=c, value=valeur)
            if cle in ("url", "url_linkedin") and valeur:
                cellule.hyperlink = valeur
                cellule.font = Font(color="0563C1", underline="single")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def exporter(chemin, offres, entreprises, contacts):
    wb = Workbook()

    ws = wb.active
    ws.title = "Offres"
    _ecrire_feuille(ws, [
        ("Nouveau", "nouveau"), ("Titre du poste", "titre"),
        ("Entreprise", "entreprise"), ("Lieu", "lieu"), ("Contrat", "contrat"),
        ("Salaire", "salaire"), ("Publiée", "date"), ("Source", "source"),
        ("Lien", "url"),
    ], offres, [10, 40, 30, 22, 14, 26, 14, 12, 60])

    ws = wb.create_sheet("Entreprises")
    _ecrire_feuille(ws, [
        ("Entreprise", "entreprise"), ("Nb d'offres", "nb_offres"),
        ("Postes recherchés", "postes"), ("Lieux", "lieux"),
        ("Sources", "sources"), ("Domaine email", "domaine"),
    ], entreprises, [30, 12, 60, 30, 18, 26])

    ws = wb.create_sheet("Contacts")
    _ecrire_feuille(ws, [
        ("Entreprise", "entreprise"), ("Nom", "nom"), ("Poste", "poste"),
        ("Profil LinkedIn", "url_linkedin"), ("Email", "email"),
        ("Statut email", "statut_email"), ("Extrait profil", "extrait"),
    ], contacts, [30, 26, 40, 55, 34, 22, 60])

    wb.save(chemin)
